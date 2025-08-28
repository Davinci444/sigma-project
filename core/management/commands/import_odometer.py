# core/management/commands/import_odometer.py
import hashlib
import os
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from fleet.models import Vehicle
from reports.models import FuelUploadLog


class Command(BaseCommand):
    help = (
        "Importa odómetros desde un Excel y actualiza Vehicle.current_odometer_km.\n"
        "Requiere hoja 'TANQUEOS' con columnas: PLACA y KILOMETRAJE (ignora mayúsculas/minúsculas)."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file_path", required=True, help="Ruta local del .xlsx a procesar")
        parser.add_argument("--sheet_name", default="TANQUEOS", help="Nombre de la hoja (por defecto TANQUEOS)")
        parser.add_argument("--user_id", type=int, default=None, help="ID del usuario que ejecuta (opcional)")

    def handle(self, *args, **opts):
        file_path = opts["file_path"]
        sheet_name = opts["sheet_name"]
        user_id = opts.get("user_id")

        if not os.path.exists(file_path):
            raise CommandError(f"Archivo no encontrado: {file_path}")

        sha256 = self._sha256_file(file_path)
        size_bytes = os.path.getsize(file_path)
        original_filename = os.path.basename(file_path)

        # Evitar reprocesos
        if FuelUploadLog.objects.filter(sha256=sha256).exists():
            self.stdout.write(self.style.WARNING(f"Ya procesado: {original_filename}"))
            return

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"No se pudo abrir el Excel: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"No existe la hoja '{sheet_name}' en el archivo.")

        ws = wb[sheet_name]

        # Leer encabezados (primera fila no vacía)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and any(cell is not None for cell in row):
                header_row = [str(c).strip() if c is not None else "" for c in row]
                break
        if not header_row:
            raise CommandError("No se encontraron encabezados en la hoja.")

        # Mapear columnas requeridas
        wanted = {"PLACA": None, "KILOMETRAJE": None}
        for idx, name in enumerate(header_row):
            up = name.upper()
            if up in wanted and wanted[up] is None:
                wanted[up] = idx

        missing = [k for k, v in wanted.items() if v is None]
        if missing:
            raise CommandError(f"Faltan columnas requeridas: {', '.join(missing)}")

        placa_idx = wanted["PLACA"]
        km_idx = wanted["KILOMETRAJE"]

        # Tomar máximo KM por placa
        max_km_by_plate: dict[str, int] = defaultdict(int)
        rows_processed = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            placa_raw = row[placa_idx] if placa_idx < len(row) else None
            km_raw = row[km_idx] if km_idx < len(row) else None
            if placa_raw is None or km_raw is None:
                continue

            try:
                placa = str(placa_raw).strip().upper()
                km = int(float(km_raw))
            except Exception:
                continue

            if not placa or km < 0:
                continue

            if km > max_km_by_plate[placa]:
                max_km_by_plate[placa] = km

            rows_processed += 1

        vehicles_updated = 0
        with transaction.atomic():
            for plate, new_km in max_km_by_plate.items():
                try:
                    v = Vehicle.objects.get(plate__iexact=plate)
                except Vehicle.DoesNotExist:
                    continue

                current = v.current_odometer_km or 0
                if new_km > current:
                    v.current_odometer_km = new_km
                    try:
                        v.odometer_status = v.OdometerStatus.VALID  # opcional
                    except Exception:
                        pass
                    v.save(update_fields=["current_odometer_km", "odometer_status"])
                    vehicles_updated += 1

            FuelUploadLog.objects.create(
                original_filename=original_filename,
                sha256=sha256,
                size_bytes=size_bytes,
                sheet_name=sheet_name,
                rows_processed=rows_processed,
                vehicles_updated=vehicles_updated,
                processed_by_id=user_id,
            )

        self.stdout.write(self.style.SUCCESS(
            f"OK: {original_filename} | filas={rows_processed} | vehiculos_actualizados={vehicles_updated}"
        ))

    def _sha256_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()